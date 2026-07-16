import os
import json
import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

def read_doc(file_path, page=1, chunk_size=8000):
    path = Path(file_path)
    if not path.exists() or not path.is_file():
        print(json.dumps({"error": f"File not found: {file_path}"}))
        return

    ext = path.suffix.lower()
    text_content = ""

    try:
        if ext == '.txt':
            with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                text_content = f.read()
        elif ext == '.csv':
            import csv
            with open(path, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                for row in reader:
                    text_content += ",".join(row) + "\n"
        elif ext == '.pdf':
            try:
                import PyPDF2
                with open(path, 'rb') as f:
                    reader = PyPDF2.PdfReader(f)
                    for p in reader.pages:
                        extracted = p.extract_text()
                        if extracted:
                            text_content += extracted + "\n"
            except ImportError:
                print(json.dumps({"error": "PyPDF2 is not installed. Please run: pip install PyPDF2"}))
                return
        elif ext in ['.doc', '.docx']:
            try:
                # Use pure Python zipfile extraction to avoid lxml/python-docx dependency issues
                with zipfile.ZipFile(path) as docx:
                    xml_content = docx.read('word/document.xml')
                    tree = ET.fromstring(xml_content)
                    for node in tree.iter():
                        if node.tag.endswith('}t') and node.text:
                            text_content += node.text + " "
            except Exception as e:
                print(json.dumps({"error": f"Failed to parse DOCX natively: {str(e)}"}))
                return
        elif ext == '.xlsx':
            try:
                import openpyxl
                # data_only=True reads cell values instead of formulas
                wb = openpyxl.load_workbook(path, data_only=True)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    text_content += f"--- Sheet: {sheet_name} ---\n"
                    for row in sheet.iter_rows(values_only=True):
                        row_strs = [str(cell) if cell is not None else "" for cell in row]
                        if any(row_strs):
                            text_content += "\t".join(row_strs) + "\n"
            except ImportError:
                print(json.dumps({"error": "openpyxl is not installed. Please run: pip install openpyxl"}))
                return
        else:
            print(json.dumps({"error": f"Unsupported file type: {ext}"}))
            return

        total_length = len(text_content)
        total_pages = max(1, (total_length + chunk_size - 1) // chunk_size)
        page = max(1, min(page, total_pages)) # Clamp page number
        
        start_idx = (page - 1) * chunk_size
        end_idx = start_idx + chunk_size

        truncated = text_content[start_idx:end_idx]
        print(json.dumps({
            "ok": True,
            "path": str(path),
            "content": truncated,
            "page": page,
            "total_pages": total_pages,
            "truncated": total_length > end_idx
        }))
    except Exception as e:
        print(json.dumps({"error": str(e)}))

if __name__ == "__main__":
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        page_num = 1
        if len(sys.argv) > 2:
            try:
                page_num = int(sys.argv[2])
            except ValueError:
                pass
        read_doc(file_path, page_num)
    else:
        print(json.dumps({"error": "Usage: doc_reader.py <file_path> [page]"}))
